#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from TestLog import log
import zipfile
from Sys_Info import SystemInfo
import time
import os


class ExcleDataSave(object):
    dirname = 'Nature(' + time.strftime('%m-%d %H%:%M', time.localtime(time.time())) + ')'
    dirpath = './Datafile/' + dirname
    documentpath = './Datafile/' + dirname + '/NatureTest.xlsx'
    log = log.log('创建文档')

    # 使用此类写数据时需要先调用此方法确认是否创建了表格
    @classmethod
    def createxcle(cls):
        if os.path.exists(cls.dirpath) is False:
            os.mkdir(cls.dirpath)
            cls.log.info("创建目录:%s" % cls.dirname)
        else:
            cls.log.info("目录%s已存在" % cls.dirpath)

        # 创建测试文档
        if os.path.exists(cls.documentpath) is False:
            excle = Workbook(cls.documentpath)
            excle.save(cls.documentpath)
            cls.log.info("创建文档:NatureTest.xlsx")
        else:
            cls.log.info("文档NatureTest.xlsx已存在")

    # 创建Cpu表格表头
    @classmethod
    def creat_cputable(cls):
        loadexcle = load_workbook(cls.documentpath)
        loadexcle.create_sheet('CpuData', 0)
        loadexcle['CpuData']['B2'] = '当前时间'
        loadexcle['CpuData']['C2'] = '当前页面'
        loadexcle['CpuData']['D2'] = '当前占用'
        loadexcle['CpuData']['E2'] = '系统核数'
        loadexcle.save(cls.documentpath)

    # 创建Mem表格表头
    @classmethod
    def creat_memtable(cls):
        loadexcle = load_workbook(cls.documentpath)
        loadexcle.create_sheet('MemData', 1)
        loadexcle['MemData']['B2'] = '当前时间'
        loadexcle['MemData']['C2'] = '当前页面'
        loadexcle['MemData']['D2'] = '系统内存(MB)'
        loadexcle['MemData']['E2'] = '应用内存(MB)'
        loadexcle.save(cls.documentpath)

    # 创建Net表格表头
    @classmethod
    def creat_nettable(cls):
        loadexcle = load_workbook(cls.documentpath)
        loadexcle.create_sheet('NetData', 2)
        # 创建表头 上行数据表头
        loadexcle['NetData']['B2'] = '系统时间'
        loadexcle['NetData']['C2'] = '上行数据(bytes)'
        loadexcle['NetData']['D2'] = '每秒消耗(kb)'
        # # # 创建表头 下行数据表头
        loadexcle['NetData']['F2'] = '系统时间'
        loadexcle['NetData']['G2'] = '下行数据(bytes)'
        loadexcle['NetData']['H2'] = '每秒消耗(kb)'
        loadexcle.save(cls.documentpath)

    @classmethod
    def creat_batterytable(cls):
        loadexcle = load_workbook(cls.documentpath)
        loadexcle.create_sheet('BatteryData', 3)
        loadexcle['BatteryData']['B2'] = '当前时间'
        loadexcle['BatteryData']['C2'] = '电池容量'
        loadexcle['BatteryData']['D2'] = '当前页面'
        loadexcle['BatteryData']['E2'] = '当前温度'
        loadexcle['BatteryData']['F2'] = '此刻耗电(mAh)'
        loadexcle['BatteryData']['G2'] = '定时耗电(mAh)'
        loadexcle.save(cls.documentpath)

    # 创建Fps表格表头
    @classmethod
    def creat_fpstable(cls):
        loadexcle = load_workbook(cls.documentpath)
        loadexcle.create_sheet('FpsData', 4)
        loadexcle['FpsData']['B2'] = '当前时间'
        loadexcle['FpsData']['C2'] = '当前页面'
        loadexcle['FpsData']['D2'] = '平均时长(ms)'
        loadexcle['FpsData']['E2'] = '掉帧率(%)'
        loadexcle.save(cls.documentpath)

    # cpudata 写入 excle
    @classmethod
    def write_cpudata(cls, row=3, cpudata=None):
        try:
            loadexcle = load_workbook(cls.documentpath)
            sheetopernew = loadexcle['CpuData']
            currenttime, currentactivity, current_percent, sys_cure = cpudata
            sheetopernew.cell(row=row, column=2, value=currenttime)
            sheetopernew.cell(row=row, column=3, value=currentactivity)
            sheetopernew.cell(row=row, column=4, value=current_percent)
            sheetopernew.cell(row=row, column=5, value=sys_cure)
            loadexcle.save(cls.documentpath)
        except zipfile.BadZipFile:
            cls.log.error("处理器Cpu数据写入出错")

    # Memdata 写入 excle
    @classmethod
    def write_memdata(cls, row=3, memdata=None):
        try:
            loadexcle = load_workbook(cls.documentpath)
            sheetopernew = loadexcle['MemData']
            currenttime, currentactivity, process_memory, sys_memory = memdata
            sheetopernew.cell(row=row, column=2, value=currenttime)
            sheetopernew.cell(row=row, column=3, value=currentactivity)
            sheetopernew.cell(row=row, column=4, value=process_memory)
            sheetopernew.cell(row=row, column=5, value=sys_memory)
            loadexcle.save(cls.documentpath)
        except zipfile.BadZipFile:
            cls.log.error("内存Mem数据写入出错")

    # 网络数据写入文件方法；txnetdata 上行 rxnetdata 下行
    @classmethod
    def write_netdata(cls, row=3, txnetdata=None, rxnetdata=None):
        try:
            loadexcle = load_workbook(cls.documentpath)
            sheetopernew = loadexcle['NetData']
            currenttime_tx, tx_data, tx_sec_data = txnetdata
            currenttime_rx, rx_data, rx_sec_data = rxnetdata
            sheetopernew.cell(row=row, column=2, value=currenttime_tx)
            sheetopernew.cell(row=row, column=3, value=tx_data)
            sheetopernew.cell(row=row, column=4, value=tx_sec_data)
            sheetopernew.cell(row=row, column=6, value=currenttime_rx)
            sheetopernew.cell(row=row, column=7, value=rx_data)
            sheetopernew.cell(row=row, column=8, value=rx_sec_data)
            loadexcle.save(cls.documentpath)
        except zipfile.BadZipFile:
            cls.log.error("流量数据写入出错")

    # battery 写入 excle
    @classmethod
    def write_batterydata(cls, row=3, batterydata=None):
        try:
            loadexcle = load_workbook(cls.documentpath)
            sheetopernew = loadexcle['BatteryData']
            currenttime, currentactivity, batterycapacity, temperature, current_power, power_persecond = batterydata
            sheetopernew.cell(row=row, column=2, value=currenttime)
            sheetopernew.cell(row=row, column=3, value=currentactivity)
            sheetopernew.cell(row=row, column=4, value=batterycapacity)
            sheetopernew.cell(row=row, column=5, value=temperature)
            sheetopernew.cell(row=row, column=6, value=current_power)
            sheetopernew.cell(row=row, column=7, value=power_persecond)
            loadexcle.save(cls.documentpath)
        except zipfile.BadZipFile:
            cls.log.error("电量数据写入出错")

    # 向 excle 里写Fps 数据方法
    @classmethod
    def write_fpsdata(cls, row=3, framesdata=None):
        try:
            loadexcle = load_workbook(cls.documentpath)
            sheetopernew = loadexcle['FpsData']
            currenttime = SystemInfo.systime()
            currentactivity = framesdata[1]
            frames_timeaverage = framesdata[0]['timeaverage']
            frames_jankyframespercent = framesdata[0]['jankyframespercent']
            sheetopernew.cell(row=row, column=2, value=currenttime)
            sheetopernew.cell(row=row, column=3, value=currentactivity)
            sheetopernew.cell(row=row, column=4, value=frames_timeaverage)
            sheetopernew.cell(row=row, column=5, value=frames_jankyframespercent)
            loadexcle.save(cls.documentpath)
        except zipfile.BadZipFile:
            cls.log.error("帧率Fps数据写入出错")


if __name__ == '__main__':
    ExcleDataSave.createxcle()
    ExcleDataSave.creat_cputable()
    ExcleDataSave.creat_memtable()
    ExcleDataSave.creat_nettable()
    ExcleDataSave.creat_batterytable()
    ExcleDataSave.creat_fpstable()
